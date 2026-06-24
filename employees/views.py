import json
import calendar
from collections import defaultdict
from calendar import monthrange
from datetime import date
from pathlib import Path

from django.conf import settings
from django.http import JsonResponse, HttpResponseBadRequest, FileResponse
from django.shortcuts import render, get_object_or_404
from django.db.models import Sum
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.urls import reverse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import Employee, Department, EmployeeStatus, AbsenceRecord, WorkDay
from .forms import EmployeeFilterForm


MONTH_NAMES_RU = {
    1: 'январь',
    2: 'февраль',
    3: 'март',
    4: 'апрель',
    5: 'май',
    6: 'июнь',
    7: 'июль',
    8: 'август',
    9: 'сентябрь',
    10: 'октябрь',
    11: 'ноябрь',
    12: 'декабрь',
}


def home(request):
    total_employees = Employee.objects.count()
    departments_count = Department.objects.count()

    status_list = []
    for status in EmployeeStatus.objects.all():
        count = Employee.objects.filter(current_status=status).count()
        status_list.append({'status': status, 'count': count})

    return render(request, 'employees/home.html', {
        'total_employees': total_employees,
        'departments_count': departments_count,
        'status_counts': status_list,
    })


def employee_list(request):
    form = EmployeeFilterForm(request.GET or None)
    employees = Employee.objects.select_related('position', 'department', 'current_status').order_by('full_name')

    search = request.GET.get('search', '').strip()
    if search:
        employees = employees.filter(full_name__icontains=search)

    if form.is_valid():
        department = form.cleaned_data.get('department')
        status = form.cleaned_data.get('status')
        if department:
            employees = employees.filter(department=department)
        if status:
            employees = employees.filter(current_status=status)

    return render(request, 'employees/employee_list.html', {
        'employees': employees,
        'form': form,
    })


def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    absence_records = AbsenceRecord.objects.filter(employee=employee).order_by('-start_date')
    statuses = EmployeeStatus.objects.all()

    return render(request, 'employees/employee_detail.html', {
        'employee': employee,
        'absence_records': absence_records,
        'statuses': statuses,
    })


@require_http_methods(['GET'])
def employee_calendar_api(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    try:
        year = int(request.GET.get('year'))
        month = int(request.GET.get('month'))
        if month < 1 or month > 12:
            raise ValueError
    except (TypeError, ValueError):
        return HttpResponseBadRequest('Некорректный месяц')

    last_day = calendar.monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last_day)

    workdays = WorkDay.objects.filter(
        employee=employee,
        date__range=(start, end)
    ).select_related('status')

    days = {}
    for item in workdays:
        days[item.date.isoformat()] = {
            'id': item.id,
            'status_id': item.status_id,
            'status_code': item.status.code if item.status else None,
            'status_name': item.status.name if item.status else None,
            'hours': item.hours,
        }

    total_hours = workdays.aggregate(total=Sum('hours'))['total'] or 0

    return JsonResponse({
        'employee_id': employee.id,
        'year': year,
        'month': month,
        'last_day': last_day,
        'days': days,
        'total_hours': total_hours,
    })


@require_http_methods(['POST'])
def employee_calendar_day_api(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return HttpResponseBadRequest('Некорректный JSON')

    day_str = payload.get('date')
    status_id = payload.get('status_id')
    hours = payload.get('hours')

    try:
        day = date.fromisoformat(day_str)
    except Exception:
        return HttpResponseBadRequest('Некорректная дата')

    sunday = day.weekday() == 6

    status = None
    if status_id not in ['', None]:
        status = get_object_or_404(EmployeeStatus, pk=status_id)

    if sunday:
        off_status = EmployeeStatus.objects.filter(name__icontains='выход').first()
        status = off_status
        hours = 0
    else:
        status_name = status.name if status else ''
        is_working = 'работ' in status_name.lower()

        if not is_working:
            hours = 0
        else:
            if hours in ['', None]:
                hours = None
            else:
                try:
                    hours = int(hours)
                except (TypeError, ValueError):
                    return HttpResponseBadRequest('Часы должны быть числом')
                if hours < 0 or hours > 16:
                    return HttpResponseBadRequest('Часы должны быть от 0 до 16')

    obj, _ = WorkDay.objects.get_or_create(employee=employee, date=day)
    obj.status = status
    obj.hours = hours
    obj.save()

    employee.current_status = status
    employee.save(update_fields=['current_status'])

    return JsonResponse({
        'ok': True,
        'date': day.isoformat(),
        'employee_status_id': employee.current_status_id,
        'employee_status_code': employee.current_status.code if employee.current_status else None,
        'employee_status_name': employee.current_status.name if employee.current_status else None,
        'status_id': obj.status_id,
        'status_code': obj.status.code if obj.status else None,
        'status_name': obj.status.name if obj.status else None,
        'hours': obj.hours,
    })


def _build_timesheet_file(year, month, force=False):
    media_dir = Path(settings.MEDIA_ROOT)
    media_dir.mkdir(parents=True, exist_ok=True)

    filename = f'tabel_t13_{year}_{month:02d}.xlsx'
    file_path = media_dir / filename
    if file_path.exists() and not force:
        return file_path, filename

    if file_path.exists() and force:
        file_path.unlink()

    days_in_month = monthrange(year, month)[1]
    employees = Employee.objects.select_related('position', 'department').order_by('full_name')

    workdays = WorkDay.objects.filter(
        employee__in=employees,
        date__year=year,
        date__month=month,
    ).select_related('status')

    data = defaultdict(dict)
    for item in workdays:
        data[item.employee_id][item.date.day] = item

    wb = Workbook()
    ws = wb.active
    ws.title = f'{month:02d}.{year}'

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    section_fill = PatternFill('solid', fgColor='EDEDED')
    end_col = 6 + days_in_month

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws['A1'] = f'Табель учета рабочего времени за {MONTH_NAMES_RU[month]}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws['A2'] = f'{month:02d}.{year}'
    ws['A2'].alignment = center

    headers = ['№', 'ФИО', 'Должность', 'Подразделение', 'Таб. номер']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for day in range(1, days_in_month + 1):
        cell = ws.cell(row=4, column=5 + day, value=day)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    cell = ws.cell(row=4, column=end_col, value='Итого часов')
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

    all_departments = list(Department.objects.order_by('name'))
    department_totals = {dept.name: 0 for dept in all_departments}
    department_totals['Без подразделения'] = 0

    row = 5
    for idx, emp in enumerate(employees, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=emp.full_name)
        ws.cell(row=row, column=3, value=emp.position.name if emp.position else '')
        ws.cell(row=row, column=4, value=emp.department.name if emp.department else '')
        ws.cell(row=row, column=5, value=idx)

        employee_total = 0

        for day in range(1, days_in_month + 1):
            current_date = date(year, month, day)
            sunday = current_date.weekday() == 6

            item = data.get(emp.id, {}).get(day)
            value = ''

            if sunday:
                value = 'В 0'
            elif item:
                status_code = item.status.code if item.status else ''
                hours_value = item.hours

                if status_code and hours_value is not None:
                    value = f'{status_code} {hours_value}'
                elif status_code:
                    if item.status and 'выход' in item.status.name.lower():
                        value = 'В 0'
                    else:
                        value = status_code
                elif hours_value is not None:
                    value = str(hours_value)

                if hours_value is not None:
                    employee_total += int(hours_value)

            cell = ws.cell(row=row, column=5 + day, value=value)
            cell.alignment = center
            cell.border = border

        ws.cell(row=row, column=end_col, value=employee_total)

        dept_name = emp.department.name if emp.department else 'Без подразделения'
        department_totals[dept_name] = department_totals.get(dept_name, 0) + employee_total

        for col in range(1, end_col + 1):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = center

        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value='Итоги по подразделениям')
    cell.font = Font(bold=True)
    cell.fill = section_fill
    cell.alignment = center
    cell.border = border

    row += 1
    ws.cell(row=row, column=1, value='Подразделение')
    ws.cell(row=row, column=2, value='Итого часов')
    for col in range(1, 3):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = center
        c.border = border

    row += 1
    for dept in all_departments:
        total = department_totals.get(dept.name, 0)
        ws.cell(row=row, column=1, value=dept.name)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2).alignment = center
        row += 1

    ws.cell(row=row, column=1, value='Без подразделения')
    ws.cell(row=row, column=2, value=department_totals.get('Без подразделения', 0))
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2).border = border
    ws.cell(row=row, column=1).alignment = center
    ws.cell(row=row, column=2).alignment = center

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    for col in range(6, end_col + 1):
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = 10
    ws.column_dimensions['F'].width = 12

    wb.save(file_path)
    return file_path, filename


def timesheet_index(request):
    year = timezone.localdate().year
    files = []
    for month in range(1, 13):
        file_path, filename = _build_timesheet_file(year, month)
        updated_at = timezone.localtime(
            timezone.datetime.fromtimestamp(file_path.stat().st_mtime, tz=timezone.get_current_timezone())
        )
        files.append({
            'month': month,
            'name': MONTH_NAMES_RU[month],
            'filename': filename,
            'url': reverse('employees:timesheet_download', args=[filename]),
            'updated_at': updated_at.strftime('%d.%m.%Y'),
        })

    return render(request, 'employees/timesheets.html', {
        'year': year,
        'files': files,
    })


@require_http_methods(['POST'])
def timesheet_refresh(request, year, month):
    file_path, filename = _build_timesheet_file(year, month, force=True)
    updated_at = timezone.localtime(
        timezone.datetime.fromtimestamp(file_path.stat().st_mtime, tz=timezone.get_current_timezone())
    )
    return JsonResponse({
        'ok': True,
        'filename': filename,
        'url': reverse('employees:timesheet_download', args=[filename]),
        'updated_at': updated_at.strftime('%d.%m.%Y'),
    })


@require_http_methods(['GET'])
def timesheet_download(request, filename):
    file_path = Path(settings.MEDIA_ROOT) / filename
    if not file_path.exists():
        return HttpResponseBadRequest('Файл не найден')

    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)


def about(request):
    return render(request, 'employees/about.html', {})